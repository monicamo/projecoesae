from main import db, app
import openpyxl
import logging
import os
import time
import pandas as pd
from datetime import datetime
from pymongo import MongoClient
from bson import ObjectId
from utils import gerar_data_referencia

class CollectionPlanilhas:
    def __init__(self):
        self.planilhas = []

    def adicionar_planilha(self, planilha):
        self.planilhas.append(planilha)

    def processar_todas_planilhas(self):
        informacoes_gerais = []

        for planilha in self.planilhas:
            informacoes_planilha = planilha.processar_planilha()
            informacoes_gerais.append(informacoes_planilha)

        return informacoes_gerais


class Planilhas(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    nome = db.Column(db.String(50), nullable=False)
    categoria = db.Column(db.String(40), nullable=False)
    console = db.Column(db.String(20), nullable=False)
    arquivo = db.Column(db.String(100), nullable=False)

    def __repr__(self):
        return '<Name %r>' % self.name

    def processar_planilha(self):
        categorias_e_funcoes = {
            'Servicos': self.processar_planilha_servicos,
            'IGPD': self.processar_planilha_igpd,
        }
        return categorias_e_funcoes.get(self.categoria, None)()

    def processar_planilha_servicos(self):
        try:
            # Carregar a planilha
            excel_arq = os.path.join(app.config['UPLOAD_EXCEL_PATH'], self.arquivo )
            wb = openpyxl.load_workbook(excel_arq)
            sheet = wb.active

            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1))]  # Assuming headers are in the first row
            header_mensal, header_12meses, header_anual = headers[1], headers[2], str(int(headers[3]))

            data_inicio_mensal, data_fim_mensal = gerar_data_referencia(header_mensal)
            data_inicio_12meses, data_fim_12meses = gerar_data_referencia(header_12meses)
            data_inicio_anual, data_fim_anual = gerar_data_referencia(header_anual)

            informacoes_processadas = []

            for row in sheet.iter_rows(min_row=2):
                # Extrair dados da linha
                instituicao = row[0].value
                mensal = row[1].value
                doze_meses = row[2].value
                anual = row[3].value

                instituicao_codigo = self.buscar_codigo_instituicao(instituicao)
                planilha_codigo = self.buscar_codigo_tipo_planilha(self.categoria)

                mensal_code = 'ML'
                dozemeses_code = '12'
                anual_code = 'AL'

                # Converter mensal para string no formato MMYY/MMYY
                mensal_str = header_mensal  # f"{mensal:%b-%y}"

                # Formatar nome do período para mensal
                codigo_mensal = f"{instituicao_codigo}{mensal_code}{planilha_codigo}"
                codigo_12meses = f"{instituicao_codigo}{dozemeses_code}{planilha_codigo}"
                codigo_anual = f"{instituicao_codigo}{anual_code}{planilha_codigo}"

                if not self.verificar_existencia_registro(instituicao, header_mensal, planilha_codigo):
                    informacoes_processadas.append({
                        'instituicao': instituicao,
                        'codigo': codigo_mensal,
                        'periodo': 'M',
                        'tipo': planilha_codigo,
                        'descricao': header_mensal,
                        'valor': mensal,
                        'data_referencia_inicial': data_inicio_mensal,
                        'data_referencia_final': data_fim_mensal,
                        'data_insercao': datetime.now().date().isoformat(),
                        'hora_insercao': datetime.now().time().isoformat()
                    })

                # Adicionar informações processadas para o período de 12 meses
                if not self.verificar_existencia_registro(instituicao, header_12meses, planilha_codigo):
                    informacoes_processadas.append({
                        'instituicao': instituicao,
                        'codigo': codigo_12meses,
                        'periodo': '12M',
                        'tipo': planilha_codigo,
                        'valor': doze_meses,
                        'descricao': header_12meses,
                        'data_referencia_inicial': data_inicio_12meses,
                        'data_referencia_final': data_fim_12meses,
                        'data_insercao': datetime.now().date().isoformat(),
                        'hora_insercao': datetime.now().time().isoformat()
                    })

                # Adicionar informações processadas para o período anual
                if not self.verificar_existencia_registro(instituicao, header_anual, planilha_codigo):
                    informacoes_processadas.append({
                        'instituicao': instituicao,
                        'codigo': codigo_anual,
                        'periodo': 'Anual',
                        'tipo': planilha_codigo,
                        'valor': anual,
                        'descricao': header_anual,
                        'data_referencia_inicial': data_inicio_anual,
                        'data_referencia_final': data_fim_anual,
                        'data_insercao': datetime.now().date().isoformat(),
                        'hora_insercao': datetime.now().time().isoformat()
                    })

            # Retornar as informações processadas
            return informacoes_processadas

        except Exception as e:
            # Registrar o erro em um arquivo de log
            logging.error(f"Erro ao processar planilha: {e}")
            # Retornar um valor vazio para indicar que a função falhou
            return []

    def buscar_codigo_tipo_planilha(self, tipo):
        # Estabelecer conexão com o MongoDB
        client = MongoClient('localhost', 27017)
        db = client['planilhas']
        colecao = db['planilhas']

        # Realizar a consulta para buscar o código da instituição pelo tipo
        planilha = colecao.find_one({'tipo': tipo})

        # Fechar a conexão com o MongoDB
        client.close()

        # Se a instituição existir, retornar o código; senão, retornar None
        if planilha:
            return planilha['codigo']
        else:
            return None

    def buscar_codigo_instituicao(self, nome):
        # Estabelecer conexão com o MongoDB
        client = MongoClient('localhost', 27017)
        db = client['planilhas']
        colecao = db['instituicoes']

        # Realizar a consulta para buscar o código da instituição pelo nome
        instituicao = colecao.find_one({'nome': nome})

        # Fechar a conexão com o MongoDB
        client.close()

        # Se a instituição existir, retornar o código; senão, retornar None
        if instituicao:
            return instituicao['codigo']
        else:
            return None

    def calcular_variacao(self, valor_atual, valor_anterior):
        if valor_anterior is None:
            return None
        return round((valor_atual - valor_anterior) / valor_anterior * 100, 2)

    def processar_planilha_igpd(self):
        # Lógica específica para planilhas do tipo 'IGPD'
        print('IGPD')
        return {'mensagem': f'Processando planilha IGPD: {self.nome}'}

    def verificar_existencia_registro(self, codigo_instituicao, codigo_periodo, codigo_planilha):
        client = MongoClient('localhost', 27017)
        db = client['planilhas']
        colecao_ativos = db['ativos']

        # Realizar a consulta na coleção
        ativo_existente = colecao_ativos.find_one({'instituicao': codigo_instituicao,
                                                   'descricao': codigo_periodo,
                                                   'codigo_planilha': codigo_planilha})

        return ativo_existente is not None


class Usuarios(db.Model):
    nickname = db.Column(db.String(8), primary_key=True)
    nome = db.Column(db.String(20), nullable=False)
    senha = db.Column(db.String(100), nullable=False)

    def __repr__(self):
        return '<Name %r>' % self.name


class ManipuladorPlanilha:
    def __init__(self, app):
        self.app = app

    def salvar_planilha(self, arquivo, nova_planilha):
        timestamp = time.time()
        caminho_arquivo = self._gerar_caminho_arquivo(nova_planilha, timestamp)
        arquivo.save(caminho_arquivo)
        return caminho_arquivo

    def ler_planilha(self, caminho_arquivo):
        return pd.read_excel(caminho_arquivo, engine='openpyxl')

    def _gerar_caminho_arquivo(self, nova_planilha, timestamp):
        nome_arquivo = f'excel{nova_planilha.id}-{timestamp}.xlsx'
        return os.path.join(self.app.config['UPLOAD_EXCEL_PATH'], nome_arquivo)


from bson import ObjectId


class Ativo:
    def __init__(self, instituicao, codigo, periodo, valor, descricao, data_inicio, data_fim, data_insercao, hora_insercao):
        self.instituicao = instituicao
        self.codigo = codigo
        self.periodo = periodo
        self.valor = valor
        self.descricao = descricao
        self.data_inicio = data_insercao
        self.data_fim = data_insercao
        self.data_insercao = data_insercao
        self.hora_insercao = hora_insercao

    @classmethod
    def get_ativo_by_id(cls, ativo_id):
        client = MongoClient('localhost', 27017)
        db = client['planilhas']
        collection = db['ativos']

        # Convertendo ativo_id para ObjectId
        ativo_id = ObjectId(ativo_id)

        ativo_data = collection.find_one({"_id": ativo_id})
        ativo = cls(
            ativo_data["instituicao"],
            ativo_data["codigo"],
            ativo_data["periodo"],
            ativo_data["valor"],
            ativo_data["descricao"],
            ativo_data["data_referencia_inicial"],
            ativo_data["data_referencia_final"],
            ativo_data["data_insercao"],
            ativo_data["hora_insercao"]
        )

        client.close()

        return ativo

